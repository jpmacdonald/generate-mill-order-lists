from dateutil.parser import parse
from pathlib import Path
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import numpy as np
import pandas as pd
import datetime
import glob
import os


def main():
    filename = 'mill-order-master.xlsx'
    pathlist = glob.glob('**/**/*.xlsx')
    data = pd.DataFrame()
    create_document(pathlist, data)
    format_document(filename)
    expand_columns(filename)


def create_document(pathlist, data):
    for path in pathlist:
        try:
            with open(path, 'rb') as file:
                print(os.path.abspath(path))
                df = pd.read_excel(
                    file,
                    sheet_name='Mill Order List',
                    engine='openpyxl',)
                if tmp := get_info(file, "Project Job #"):
                    jobNum = tmp
                else:
                    jobNum = f'Error: {path}'
                if tmp := get_info(file, "Project Name"):
                    projName = tmp
        except:
            print(f'There was a problem with file: {path}')
            continue

        df = df.dropna(how='all')
        df, df.columns = df[4:], df.iloc[3]
        # trim extra rows that are irrelevant
        try:
            df = df[df['Mill Order'] != 0]
            df = df[df['Mill Order'].notnull()]
            df = df.loc[:, pd.notnull(df.columns)]

            # trim not needed columns (any after 'Actual', i.e. Delivery)
            df = df.loc[:, :'Actual']

            # renaming columns in case of spelling errors
            df.columns = ['Mill Order', 'Date Assigned', 'Description',
                          'Mechanic', 'Date Due', 'Actual']

            df['Job Number'] = jobNum
            df = df[['Job Number', 'Mill Order', 'Mechanic', 'Description', 'Date Assigned',
                     'Date Due', 'Actual']]
            data = data.append(df, ignore_index=True)
        except:
            print(f"Problem locating 'Mill Order' sheet: {path}")
            continue

    data = data.sort_values(by=['Job Number', 'Mill Order'])
    data = data.replace(np.nan, '', regex=True)
    data = data.replace(0, '', regex=True)
    data['Mechanic'] = data['Mechanic'].str.upper()

    data = data[~(data['Description'] == '')]
    data['Description'] = data['Description'].str.strip()
    data['Description'] = data['Description'].str.upper()

    timestr = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f'{timestr}_mill-order-master.xlsx'
    writer = ExcelWriter(filename)
    data.to_excel(writer, 'Master', index=False)
    writer.save()


def format_document(filename):
    # need to normalize date columns before formatting
    format_dates('%b-%d-%Y', filename)
    expand_columns(filename)


def format_dates(date_format, filename):
    date_format = '%b-%d-%Y'
    wb = load_workbook(filename=filename)
    alignment = Alignment(horizontal='center')
    ws = wb.active
    for row in ws[2: ws.max_row]:        # skip the header
        cell = row[4]                   # column E
        if isinstance(cell.value, datetime.datetime):
            cell.value = cell.value.strftime(date_format)
            cell.alignment = alignment
        cell = row[5]                   # column F
        if isinstance(cell.value, datetime.datetime):
            cell.value = cell.value.strftime(date_format)
            cell.alignment = alignment
        cell = row[6]                   # column G
        if isinstance(cell.value, datetime.datetime):
            cell.value = cell.value.strftime(date_format)
            cell.alignment = alignment
    wb.save(filename)


def expand_columns(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell.coordinate in ws.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(filename)


def is_date(string, fuzzy=False):
    try:
        parse(string, fuzzy=fuzzy)
        return True
    except ValueError:
        return False


def get_info(file, string):
    wb = load_workbook(file, read_only=True)
    ws = wb['Info Sheet']

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == string:
                return ws.cell(row=cell.row, column=2).value
    return None


main()
